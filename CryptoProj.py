import time

start = time.time()

import openpyxl
import requests
import xlsxwriter
import sys
import re
import matplotlib.pyplot as plt
from matplotlib.ticker import ScalarFormatter, FormatStrFormatter
import numpy as np
import matplotlib.patches as mpatches
import matplotlib.dates as mdates
from datetime import date
import datetime



def calcSSMA(list):
    sum = 0
    if (len(list) < 50):
        return '-'
    for elem in list:
        sum = sum + elem
    res = sum / (len(list))
    res = round(res,9)
    return str(res)

def calcLSMA(list):
    sum = 0
    if (len(list) < 200):
        return '-'
    for elem in list:
        sum = sum + elem
    res = sum / (len(list))
    res = round(res,9)
    return str(res)

def calcSEMA(x, list, idx, nS, p):
    if (len(list) < nS):
        return '-'
    k = 2/(nS+1)
    prev = 0
    if list[idx] != '-':
        prev = float(list[idx])
    else:
        prev = float(p)
    res = k * prev + (1-k) * x
    pSema = res
    res = round(res,9)
    return str(res)

def calcLEMA(x, list, idx, nL, p):
    if (len(list) < nL):
        return '-'
    k = 2/(nL+1)
    prev = 0
    if list[idx] != '-':
        prev = float(list[idx])
    else:
        prev = float(p)
    res = k * prev + (1-k) * x
    pLema = res
    res = round(res,9)
    return str(res)

def infCalc(values):
    nS = 50
    nL = 200
    shortArray = []
    longArray = []
    ssmaList = []
    lsmaList = []
    semaList = []
    lemaList = []
    
    pSema = 0
    pLema = 0
    pSema2 = 0
    pLema2 = 0
    
    i = -1
    for val in values:
        elem = float(val)
        if len(shortArray) < nS:
            shortArray.append(elem)
        else:
            shortArray.pop(0)
            shortArray.append(elem)
        
        if len(longArray) < nL:
            longArray.append(elem)
        else:
            longArray.pop(0)
            longArray.append(elem)
    
        ssma = calcSSMA(shortArray)
        ssmaList.append(ssma)
        
        lsma = calcLSMA(longArray)
        lsmaList.append(lsma)
        
        pSema2 = pSema
        pLema2 = pLema
        pSema = ssma
        pLema = lsma
        
        
        sema = calcSEMA(elem, semaList, i, nS, pSema2)
        semaList.append(sema)
        
        lema = calcLEMA(elem, lemaList, i, nL, pLema2)
        lemaList.append(lema)
        

    i = i + 1
    return (ssmaList, lsmaList, semaList, lemaList)


def rsiCalcs14(values): # s14
    changeList = []
    gains = []
    losses = []
    recglist = []
    recllist = []
    aveGainList = []
    aveLossList = []
    rsList = []
    rsiList = []
    changeList.append('-')
    aveGainList.append('-')
    aveLossList.append('-')
    rsList.append('-')
    rsiList.append('-')
    gains.append(0)
    losses.append(0)
    idx = 1
    while (idx < 14):
        x = values[idx] - values[idx-1]
        x = round(x,8)
        changeList.append(x)
        aveGainList.append('-')
        aveLossList.append('-')
        rsList.append('-')
        rsiList.append('-')
        if x > 0:
            gains.append(x)
            losses.append(0)
            recglist.append(x)
            recllist.append(0)
        
        else:
            gains.append(0)
            losses.append(x)
            recglist.append(0)
            recllist.append(x)
        idx = idx + 1


    x = values[idx] - values[idx-1]
    x = round(x,8)
    changeList.append(x)
    if x > 0:
        gains.append(x)
        losses.append(0)
        recglist.append(x)
        recllist.append(0)
    else:
        gains.append(0)
        losses.append(x)
        recglist.append(0)
        recllist.append(x)

    gsum = 0
    for i in recglist:
        gsum = i + gsum
    aveGainList.append(gsum/14)
    lsum = 0
    for i in recllist:
        lsum = lsum - i
    aveLossList.append(lsum/14)
    if aveLossList[-1] == 0:
        rsList.append('Inf')
        rsiList.append(100)
    else:
        rsList.append(aveGainList[-1]/aveLossList[-1])
        rsiList.append(100 - (100 / (1 + rsList[-1])))
    idx = idx + 1


    while (idx < len(values)):
        x = values[idx] - values[idx-1]
        x = round(x,8)
        changeList.append(x)
        if x > 0:
            gains.append(x)
            losses.append(0)
            recglist.append(x)
            recllist.append(0)
        else:
            gains.append(0)
            losses.append(x)
            recglist.append(0)
            recllist.append(x)
        recglist.pop(0)
        recllist.pop(0)
        gsum = 0
        for i in recglist:
            gsum = i + gsum
        aveGainList.append(gsum/14)
        lsum = 0
        for i in recllist:
            lsum = lsum - i
        aveLossList.append(lsum/14)
        #aveGainList.append((aveGainList[-1]*13 + gains[-1]) / 14)
        #aveLossList.append((aveLossList[-1]*13 - losses[-1]) / 14)
        if aveLossList[-1] == 0:
            rsList.append('Inf')
            rsiList.append(100)
        else:
            rsList.append(aveGainList[-1]/aveLossList[-1])
            rsiList.append(100 - (100 / (1 + rsList[-1])))
        idx = idx + 1
    return (changeList, gains, losses, aveGainList, aveLossList, rsList, rsiList)




def rsiCalcw14(values): # w 14
    changeList = []
    gains = []
    losses = []
    recglist = []
    recllist = []
    aveGainList = []
    aveLossList = []
    rsList = []
    rsiList = []
    weights = list(range(1, 15))
    changeList.append('-')
    aveGainList.append('-')
    aveLossList.append('-')
    rsList.append('-')
    rsiList.append('-')
    gains.append(0)
    losses.append(0)
    idx = 1
    summ = (14*(14+1))/2
    while (idx < 14):
        x = values[idx] - values[idx-1]
        x = round(x,8)
        changeList.append(x)
        aveGainList.append('-')
        aveLossList.append('-')
        rsList.append('-')
        rsiList.append('-')
        if x > 0:
            gains.append(x)
            losses.append(0)
            recglist.append(x)
            recllist.append(0)
        
        else:
            gains.append(0)
            losses.append(x)
            recglist.append(0)
            recllist.append(x)
        idx = idx + 1
    
    
    x = values[idx] - values[idx-1]
    x = round(x,8)
    changeList.append(x)
    if x > 0:
        gains.append(x)
        losses.append(0)
        recglist.append(x)
        recllist.append(0)
    else:
        gains.append(0)
        losses.append(x)
        recglist.append(0)
        recllist.append(x)
    widx = 0
    gsum = 0
    for i in recglist:
        gsum = (i * weights[widx]) + gsum
        widx = widx+1
    aveGainList.append(gsum/summ)
    lsum = 0
    widx = 0
    for i in recllist:
        lsum = lsum - (i * weights[widx])
        widx = widx+1
    aveLossList.append(lsum/summ)
    if aveLossList[-1] == 0:
        rsList.append('Inf')
        rsiList.append(100)
    else:
        rsList.append(aveGainList[-1]/aveLossList[-1])
        rsiList.append(100 - (100 / (1 + rsList[-1])))
    idx = idx + 1
    
    
    while (idx < len(values)):
        x = values[idx] - values[idx-1]
        x = round(x,8)
        changeList.append(x)
        if x > 0:
            gains.append(x)
            losses.append(0)
            recglist.append(x)
            recllist.append(0)
        else:
            gains.append(0)
            losses.append(x)
            recglist.append(0)
            recllist.append(x)
        recglist.pop(0)
        recllist.pop(0)
        widx = 0
        gsum = 0
        for i in recglist:
            gsum = (i * weights[widx]) + gsum
            widx = widx+1
        aveGainList.append(gsum/summ)
        lsum = 0
        widx = 0
        for i in recllist:
            lsum = lsum - (i * weights[widx])
            widx = widx + 1
        aveLossList.append(lsum/summ)
        #aveGainList.append((aveGainList[-1]*13 + gains[-1]) / 14)
        #aveLossList.append((aveLossList[-1]*13 - losses[-1]) / 14)
        if aveLossList[-1] == 0:
            rsList.append('Inf')
            rsiList.append(100)
        else:
            rsList.append(aveGainList[-1]/aveLossList[-1])
            rsiList.append(100 - (100 / (1 + rsList[-1])))
        idx = idx + 1
    return (aveGainList, aveLossList, rsList, rsiList)

def rsiCalcs28(values): # s 28
    changeList = []
    gains = []
    losses = []
    recglist = []
    recllist = []
    aveGainList = []
    aveLossList = []
    rsList = []
    rsiList = []
    changeList.append('-')
    aveGainList.append('-')
    aveLossList.append('-')
    rsList.append('-')
    rsiList.append('-')
    gains.append(0)
    losses.append(0)
    idx = 1
    while (idx < 28):
        x = values[idx] - values[idx-1]
        x = round(x,8)
        changeList.append(x)
        aveGainList.append('-')
        aveLossList.append('-')
        rsList.append('-')
        rsiList.append('-')
        if x > 0:
            gains.append(x)
            losses.append(0)
            recglist.append(x)
            recllist.append(0)
        
        else:
            gains.append(0)
            losses.append(x)
            recglist.append(0)
            recllist.append(x)
        idx = idx + 1
    
    
    x = values[idx] - values[idx-1]
    x = round(x,8)
    changeList.append(x)
    if x > 0:
        gains.append(x)
        losses.append(0)
        recglist.append(x)
        recllist.append(0)
    else:
        gains.append(0)
        losses.append(x)
        recglist.append(0)
        recllist.append(x)
    
    gsum = 0
    for i in recglist:
        gsum = i + gsum
    aveGainList.append(gsum/28)
    lsum = 0
    for i in recllist:
        lsum = lsum - i
    aveLossList.append(lsum/28)
    if aveLossList[-1] == 0:
        rsList.append('Inf')
        rsiList.append(100)
    else:
        rsList.append(aveGainList[-1]/aveLossList[-1])
        rsiList.append(100 - (100 / (1 + rsList[-1])))
    idx = idx + 1
    
    
    while (idx < len(values)):
        x = values[idx] - values[idx-1]
        x = round(x,8)
        changeList.append(x)
        if x > 0:
            gains.append(x)
            losses.append(0)
            recglist.append(x)
            recllist.append(0)
        else:
            gains.append(0)
            losses.append(x)
            recglist.append(0)
            recllist.append(x)
        recglist.pop(0)
        recllist.pop(0)
        gsum = 0
        for i in recglist:
            gsum = i + gsum
        aveGainList.append(gsum/28)
        lsum = 0
        for i in recllist:
            lsum = lsum - i
        aveLossList.append(lsum/28)
        #aveGainList.append((aveGainList[-1]*13 + gains[-1]) / 14)
        #aveLossList.append((aveLossList[-1]*13 - losses[-1]) / 14)
        if aveLossList[-1] == 0:
            rsList.append('Inf')
            rsiList.append(100)
        else:
            rsList.append(aveGainList[-1]/aveLossList[-1])
            rsiList.append(100 - (100 / (1 + rsList[-1])))
        idx = idx + 1
    return (aveGainList, aveLossList, rsList, rsiList)




def rsiCalcw28(values): # w 28
    changeList = []
    gains = []
    losses = []
    recglist = []
    recllist = []
    aveGainList = []
    aveLossList = []
    rsList = []
    rsiList = []
    weights = list(range(1, 29))
    changeList.append('-')
    aveGainList.append('-')
    aveLossList.append('-')
    rsList.append('-')
    rsiList.append('-')
    gains.append(0)
    losses.append(0)
    idx = 1
    summ = (28*(28+1))/2
    while (idx < 28):
        x = values[idx] - values[idx-1]
        x = round(x,8)
        changeList.append(x)
        aveGainList.append('-')
        aveLossList.append('-')
        rsList.append('-')
        rsiList.append('-')
        if x > 0:
            gains.append(x)
            losses.append(0)
            recglist.append(x)
            recllist.append(0)
        
        else:
            gains.append(0)
            losses.append(x)
            recglist.append(0)
            recllist.append(x)
        idx = idx + 1
    
    
    x = values[idx] - values[idx-1]
    x = round(x,8)
    changeList.append(x)
    if x > 0:
        gains.append(x)
        losses.append(0)
        recglist.append(x)
        recllist.append(0)
    else:
        gains.append(0)
        losses.append(x)
        recglist.append(0)
        recllist.append(x)
    widx = 0
    gsum = 0
    for i in recglist:
        gsum = (i * weights[widx]) + gsum
        widx = widx+1
    aveGainList.append(gsum/summ)
    lsum = 0
    widx = 0
    for i in recllist:
        lsum = lsum - (i * weights[widx])
        widx = widx+1
    aveLossList.append(lsum/summ)
    if aveLossList[-1] == 0:
        rsList.append('Inf')
        rsiList.append(100)
    else:
        rsList.append(aveGainList[-1]/aveLossList[-1])
        rsiList.append(100 - (100 / (1 + rsList[-1])))
    idx = idx + 1
    
    
    while (idx < len(values)):
        x = values[idx] - values[idx-1]
        x = round(x,8)
        changeList.append(x)
        if x > 0:
            gains.append(x)
            losses.append(0)
            recglist.append(x)
            recllist.append(0)
        else:
            gains.append(0)
            losses.append(x)
            recglist.append(0)
            recllist.append(x)
        recglist.pop(0)
        recllist.pop(0)
        widx = 0
        gsum = 0
        for i in recglist:
            gsum = (i * weights[widx]) + gsum
            widx = widx+1
        aveGainList.append(gsum/summ)
        lsum = 0
        widx = 0
        for i in recllist:
            lsum = lsum - (i * weights[widx])
            widx = widx + 1
        aveLossList.append(lsum/summ)
        #aveGainList.append((aveGainList[-1]*13 + gains[-1]) / 14)
        #aveLossList.append((aveLossList[-1]*13 - losses[-1]) / 14)
        if aveLossList[-1] == 0:
            rsList.append('Inf')
            rsiList.append(100)
        else:
            rsList.append(aveGainList[-1]/aveLossList[-1])
            rsiList.append(100 - (100 / (1 + rsList[-1])))
        idx = idx + 1
    return (aveGainList, aveLossList, rsList, rsiList)


workbook = xlsxwriter.Workbook('CyrtoRes.xlsx')

worksheet1 = workbook.add_worksheet('Bitcoin')
worksheet2 = workbook.add_worksheet('Ripple')
worksheet3 = workbook.add_worksheet('Ethereum')
#''''
worksheet4 = workbook.add_worksheet('Augur')
worksheet5 = workbook.add_worksheet('Eos')
worksheet6 = workbook.add_worksheet('Litecoin')
worksheet7 = workbook.add_worksheet('Stellar')
worksheet8 = workbook.add_worksheet('Cardano')
worksheet9 = workbook.add_worksheet('Iota')
worksheet10 = workbook.add_worksheet('Tron')
worksheet11 = workbook.add_worksheet('Neo')
worksheet12 = workbook.add_worksheet('Ethereum Classic')
worksheet13 = workbook.add_worksheet('Zcash')
worksheet14 = workbook.add_worksheet('Qtum')
worksheet15 = workbook.add_worksheet('Bitcoin Cash')
#'''

def writesheet(wrksht):
    bold = workbook.add_format({'bold':True})
    wrksht.set_column('A:A', 15)
    wrksht.set_column('B:B', 20)
    wrksht.set_column('C:C', 15)
    wrksht.set_column('D:D', 15)
    wrksht.set_column('E:E', 15)
    wrksht.set_column('F:F', 15)
    wrksht.set_column('G:G', 20)
    wrksht.set_column('H:H', 15)
    wrksht.set_column('I:I', 15)
    wrksht.set_column('J:J', 15)
    wrksht.set_column('K:K', 15)
    wrksht.set_column('L:L', 15)
    wrksht.set_column('M:M', 15)
    wrksht.set_column('N:N', 15)
    wrksht.set_column('O:O', 25)
    wrksht.set_column('P:P', 25)
    wrksht.set_column('Q:Q', 25)
    wrksht.set_column('R:R', 25)
    wrksht.set_column('S:S', 28)
    wrksht.set_column('T:T', 28)
    wrksht.set_column('U:S', 25)
    wrksht.set_column('V:V', 25)
    wrksht.set_column('W:W', 25)
    wrksht.set_column('X:X', 25)
    wrksht.set_column('Y:Y', 25)
    wrksht.set_column('Z:Z', 25)
    wrksht.set_column('AA:AA', 28)
    wrksht.set_column('AB:AB', 28)
    wrksht.set_column('AC:AC', 25)
    wrksht.set_column('AD:AD', 25)

    wrksht.write('A1','DATE',bold)
    wrksht.write('B1','Open',bold)
    wrksht.write('C1', 'High', bold)
    wrksht.write('D1', 'Low', bold)
    wrksht.write('E1', 'Close', bold)
    wrksht.write('F1', 'Volume', bold)
    wrksht.write('G1', 'Market Cap', bold)
    wrksht.write('H1', 'SMA 50', bold)
    wrksht.write('I1', 'SMA 200', bold)
    wrksht.write('J1', 'EMA 50', bold)
    wrksht.write('K1', 'EMA 200', bold)
    wrksht.write('L1', 'Close Change', bold)
    wrksht.write('M1', 'Gain', bold)
    wrksht.write('N1', 'Loss', bold)
    wrksht.write('O1', '14-day simple Average Gain', bold)
    wrksht.write('P1', '14-day simple Average Loss', bold)
    wrksht.write('Q1', '14-day simple RS', bold)
    wrksht.write('R1', '14-day simple RSI', bold)
    wrksht.write('S1', '14-day weighted Average Gain', bold)
    wrksht.write('T1', '14-day weighted Average Loss', bold)
    wrksht.write('U1', '14-day weighted RS', bold)
    wrksht.write('V1', '14-day weighted RSI', bold)
    wrksht.write('W1', '28-day simple Average Gain', bold)
    wrksht.write('X1', '28-day simple Average Loss', bold)
    wrksht.write('Y1', '28-day simple RS', bold)
    wrksht.write('Z1', '28-day simple RSI', bold)
    wrksht.write('AA1', '28-day weighted Average Gain', bold)
    wrksht.write('AB1', '28-day weighted Average Loss', bold)
    wrksht.write('AC1', '28-day weighted RS', bold)
    wrksht.write('AD1', '28-day weighted RSI', bold)

writesheet(worksheet1)
writesheet(worksheet2)
writesheet(worksheet3)
#''''
writesheet(worksheet4)
writesheet(worksheet5)
writesheet(worksheet6)
writesheet(worksheet7)
writesheet(worksheet8)
writesheet(worksheet9)
writesheet(worksheet10)
writesheet(worksheet11)
writesheet(worksheet12)
writesheet(worksheet13)
writesheet(worksheet14)
writesheet(worksheet15)
#'''
wb = openpyxl.load_workbook('CoinMarket.xlsx')

sheet1 = wb.get_sheet_by_name('bitcoin')
sheet2 = wb.get_sheet_by_name('ripple')
sheet3 = wb.get_sheet_by_name('ethereum')
#''''
sheet4 = wb.get_sheet_by_name('augur')
sheet5 = wb.get_sheet_by_name('eos')
sheet6 = wb.get_sheet_by_name('litecoin')
sheet7 = wb.get_sheet_by_name('stellar')
sheet8 = wb.get_sheet_by_name('cardano')
sheet9 = wb.get_sheet_by_name('iota')
sheet10 = wb.get_sheet_by_name('tron')
sheet11 = wb.get_sheet_by_name('ethereum classic')
sheet12 = wb.get_sheet_by_name('zcash')
sheet13 = wb.get_sheet_by_name('qtum')
sheet14 = wb.get_sheet_by_name('neo')
sheet15 = wb.get_sheet_by_name('bitcoincash')
#'''

formatd = workbook.add_format({'num_format': 'mmm d yyyy'})

def writeData(sheet, wrksht):
    count = 1
    dates = []
    openList = []
    highList = []
    lowList = []
    closeList = []
    volList = []
    mcList = []
    ssmaList = []
    lsmaList = []
    semaList = []
    lemaList = []

    for i in range (2,sheet.max_row+1):
        dates.append(sheet['A' + str(i)].value)
        openList.append(sheet['B' + str(i)].value)
        highList.append(sheet['C' + str(i)].value)
        lowList.append(sheet['D' + str(i)].value)
        closeList.append(sheet['E' + str(i)].value)
        volList.append(sheet['F' + str(i)].value)
        mcList.append(sheet['G' + str(i)].value)

    dates.reverse()
    openList.reverse()
    highList.reverse()
    lowList.reverse()
    closeList.reverse()
    volList.reverse()
    mcList.reverse()

    ssmaList, lsmaList, semaList, lemaList = infCalc(closeList)
    changeList, gains, losses, aveGainList, aveLossList, rsList, rsiList = rsiCalcs14(closeList)
    aveGainList2, aveLossList2, rsList2, rsiList2 = rsiCalcw14(closeList)
    aveGainList3, aveLossList3, rsList3, rsiList3 = rsiCalcs28(closeList)
    aveGainList4, aveLossList4, rsList4, rsiList4 = rsiCalcw28(closeList)

    idx = 0

    for i in range (2,sheet.max_row+1):
        wrksht.write(count,0, dates[idx], formatd)
        wrksht.write(count,1, openList[idx])
        wrksht.write(count,2, highList[idx])
        wrksht.write(count,3, lowList[idx])
        wrksht.write(count,4, closeList[idx])
        wrksht.write(count,5, volList[idx])
        if type(volList[idx]) is str:
            volList[idx] = 0
        else:
            volList[idx] = int(volList[idx])
        wrksht.write(count,6, mcList[idx])
        wrksht.write(count,7, ssmaList[idx])
        wrksht.write(count,8, lsmaList[idx])
        wrksht.write(count,9, semaList[idx])
        wrksht.write(count,10, lemaList[idx])
        wrksht.write(count,11, changeList[idx])
        if (gains[idx] == 0):
            wrksht.write(count,12, '-')
        else:
            wrksht.write(count,12, gains[idx])
        if (losses[idx] == 0):
            wrksht.write(count,13, '-')
        else:
            wrksht.write(count,13, -1 * losses[idx])
        wrksht.write(count,14, aveGainList[idx])
        wrksht.write(count,15, aveLossList[idx])
        wrksht.write(count,16, rsList[idx])
        wrksht.write(count,17, rsiList[idx])
        if rsiList[idx] == '-':
                rsiList[idx] = 0
        else:
            rsiList[idx] = int(rsiList[idx])
        wrksht.write(count,18, aveGainList2[idx])
        wrksht.write(count,19, aveLossList2[idx])
        wrksht.write(count,20, rsList2[idx])
        wrksht.write(count,21, rsiList2[idx])
        if rsiList2[idx] == '-':
                rsiList2[idx] = 0
        else:
            rsiList2[idx] = int(rsiList2[idx])
        wrksht.write(count,22, aveGainList3[idx])
        wrksht.write(count,23, aveLossList3[idx])
        wrksht.write(count,24, rsList3[idx])
        wrksht.write(count,25, rsiList3[idx])
        if rsiList3[idx] == '-':
                rsiList3[idx] = 0
        else:
            rsiList3[idx] = int(rsiList3[idx])
        wrksht.write(count,26, aveGainList4[idx])
        wrksht.write(count,27, aveLossList4[idx])
        wrksht.write(count,28, rsList4[idx])
        wrksht.write(count,29, rsiList4[idx])
        if rsiList4[idx] == '-':
            rsiList4[idx] = 0
        else:
            rsiList4[idx] = int(rsiList4[idx])
        idx = idx + 1
        count = count + 1
    return (dates, closeList, volList, rsiList, rsiList2, rsiList3, rsiList4)

d1,c1,v1, r11, r12, r13, r14 = writeData(sheet1, worksheet1)
d1,c1,v1, r11, r12, r13, r14 = writeData(sheet2, worksheet2)
d1,c1,v1, r11, r12, r13, r14 = writeData(sheet3, worksheet3)
#''''
writeData(sheet4, worksheet4)
writeData(sheet5, worksheet5)
writeData(sheet6, worksheet6)
writeData(sheet7, worksheet7)
writeData(sheet8, worksheet8)
writeData(sheet9, worksheet9)
writeData(sheet10, worksheet10)
writeData(sheet11, worksheet11)
writeData(sheet12, worksheet12)
writeData(sheet13, worksheet13)
writeData(sheet14, worksheet14)
writeData(sheet15, worksheet15)
#'''
workbook.close()


d = datetime.datetime(2017,1,1)
pidx = d1.index(d)
d1 = d1[pidx:len(d1)]
c1 = c1[pidx:len(c1)]
v1 = v1[pidx:len(v1)]
r11 = r11[pidx:len(r11)]
r12 = r12[pidx:len(r12)]
r13 = r13[pidx:len(r13)]
r14 = r14[pidx:len(r14)]



fig, ax1 = plt.subplots()
plt.xticks(rotation=90)
fig.suptitle('Ethereum: Weighted RSIs', fontsize=20)


myFmt = mdates.DateFormatter('%Y-%m')
t = d1
s1 = c1
lw = 0.6
ax1.plot(t, s1, 'g-', linewidth = lw, alpha = 0.7)
ax1.set_xlabel('Time')
ax1.set_ylabel('Close Prices', color='k')
ax1.tick_params('y', colors='k')

ax2 = ax1.twinx()
s2 = v1
ax2.plot(t, s2, 'k-', linewidth = lw, alpha = 0.7)
ax2.set_ylabel('Volumes', color='k')
ax2.tick_params('y', colors='k')


ax3 = ax1.twinx()
#ax3.plot(t, r11, 'r-', linewidth = lw, alpha = 0.7)
ax3.plot(t, r12, 'r-', linewidth = lw, alpha = 0.7)
#ax3.plot(t, r13, 'b-', linewidth = lw, alpha = 0.7)
ax3.plot(t, r14, 'b-', linewidth = lw, alpha = 0.7)

ax3.set_ylabel('Relative Strength Index', color='k')
ax3.tick_params('y', colors='k')
ax3.spines['right'].set_position(('outward', 60))
ax1.xaxis.set_major_locator(mdates.MonthLocator(interval=1))
ax1.xaxis.set_major_formatter(myFmt)


fig.tight_layout()

patch1 = mpatches.Patch(color = 'green', label = 'Close Prices')
patch2 = mpatches.Patch(color = 'black', label = 'Volumes')
#patch3 = mpatches.Patch(color = 'red', label = '14-day Simple RSI')
patch3 = mpatches.Patch(color = 'red', label = '14-day Weighted RSI')
#patch4 = mpatches.Patch(color = 'blue', label = '28-day Simple RSI')
patch4 = mpatches.Patch(color = 'blue', label = '28-day Weighted RSI')
plt.legend(handles=[patch1, patch2, patch3, patch4], bbox_to_anchor=(0.5,-0.2), fancybox=True, shadow=True, loc = 'upper center', ncol = 5)

plt.show()
end = time.time()
elapsed = end - start
print("This program took " + str(round(elapsed,2)) + " seconds to run")
