import openpyxl
import requests
import xlsxwriter
import time
import sys
import re
from urllib.request import urlopen
try:
    import urllib.request as urllib2
except ImportError:
    import urllib2


start = time.time()



def calcSSMA(list):
    sum = 0
    if (len(list) < 50):
        return '-'
    for elem in list:
        sum = sum + elem
    res = sum / (len(list))
    res = round(res,9)
    return str(res)

def calcLSMA(list):
    sum = 0
    if (len(list) < 200):
        return '-'
    for elem in list:
        sum = sum + elem
    res = sum / (len(list))
    res = round(res,9)
    return str(res)

def calcSEMA(x, list, idx, nS, p):
    if (len(list) < nS):
        return '-'
    k = 2/(nS+1)
    prev = 0
    if list[idx] != '-':
        prev = float(list[idx])
    else:
        prev = float(p)
    res = k * prev + (1-k) * x
    pSema = res
    res = round(res,9)
    return str(res)

def calcLEMA(x, list, idx, nL, p):
    if (len(list) < nL):
        return '-'
    k = 2/(nL+1)
    prev = 0
    if list[idx] != '-':
        prev = float(list[idx])
    else:
        prev = float(p)
    res = k * prev + (1-k) * x
    pLema = res
    res = round(res,9)
    return str(res)

def infCalc(values):
    nS = 50
    nL = 200
    shortArray = []
    longArray = []
    ssmaList = []
    lsmaList = []
    semaList = []
    lemaList = []
    
    pSema = 0
    pLema = 0
    pSema2 = 0
    pLema2 = 0
    
    i = -1
    for val in values:
        elem = float(val)
        if len(shortArray) < nS:
            shortArray.append(elem)
        else:
            shortArray.pop(0)
            shortArray.append(elem)
        
        if len(longArray) < nL:
            longArray.append(elem)
        else:
            longArray.pop(0)
            longArray.append(elem)
    
        ssma = calcSSMA(shortArray)
        ssmaList.append(ssma)
        
        lsma = calcLSMA(longArray)
        lsmaList.append(lsma)
        
        pSema2 = pSema
        pLema2 = pLema
        pSema = ssma
        pLema = lsma
        
        
        sema = calcSEMA(elem, semaList, i, nS, pSema2)
        semaList.append(sema)
        
        lema = calcLEMA(elem, lemaList, i, nL, pLema2)
        lemaList.append(lema)
        

    i = i + 1
    return (ssmaList, lsmaList, semaList, lemaList)


def rsiCalc(values):
    changeList = []
    gains = []
    losses = []
    aveGainList = []
    aveLossList = []
    rsList = []
    rsiList = []
    changeList.append('-')
    aveGainList.append('-')
    aveLossList.append('-')
    rsList.append('-')
    rsiList.append('-')
    gains.append(0)
    losses.append(0)

    idx = 1
    while (idx < 14):
        x = values[idx] - values[idx-1]
        x = round(x,4)
        changeList.append(x)
        aveGainList.append('-')
        aveLossList.append('-')
        rsList.append('-')
        rsiList.append('-')
        if x > 0:
            gains.append(x)
            losses.append(0)
        else:
            gains.append(0)
            losses.append(x)
        idx = idx + 1


    x = values[idx] - values[idx-1]
    x = round(x,4)
    changeList.append(x)
    if x > 0:
        gains.append(x)
        losses.append(0)
    else:
        gains.append(0)
        losses.append(x)
    gsum = 0
    for i in gains:
        gsum = i + gsum
    aveGainList.append(gsum/14)
    lsum = 0
    for i in losses:
        lsum = lsum - i
    aveLossList.append(lsum/14)
    rsList.append(aveGainList[-1]/aveLossList[-1])
    rsiList.append(100 - (100 / (1 + rsList[-1])))
    idx = idx + 1


    while (idx < len(values)):
        x = values[idx] - values[idx-1]
        x = round(x,4)
        changeList.append(x)
        if x > 0:
            gains.append(x)
            losses.append(0)
        else:
            gains.append(0)
            losses.append(x)
        aveGainList.append((aveGainList[-1]*13 + gains[-1]) / 14)
        aveLossList.append((aveLossList[-1]*13 - losses[-1]) / 14)
        rsList.append(aveGainList[-1]/aveLossList[-1])
        rsiList.append(100 - (100 / (1 + rsList[-1])))
        idx = idx + 1
    return (changeList, gains, losses, aveGainList, aveLossList, rsList, rsiList)





workbook = xlsxwriter.Workbook('CyrtoRes.xlsx')
worksheet = workbook.add_worksheet()
bold = workbook.add_format({'bold':True})
worksheet.set_column('A:A', 15)
worksheet.set_column('B:B', 20)
worksheet.set_column('C:C', 15)
worksheet.set_column('D:D', 15)
worksheet.set_column('E:E', 15)
worksheet.set_column('F:F', 15)
worksheet.set_column('G:G', 15)
worksheet.set_column('H:H', 20)
worksheet.set_column('I:I', 15)
worksheet.set_column('J:J', 15)
worksheet.set_column('K:K', 15)
worksheet.set_column('L:L', 15)
worksheet.set_column('M:M', 15)
worksheet.set_column('N:N', 15)
worksheet.set_column('O:O', 15)
worksheet.set_column('P:P', 20)
worksheet.set_column('Q:Q', 20)
worksheet.set_column('R:R', 20)
worksheet.set_column('S:S', 20)






worksheet.write('A1','DATE',bold)
worksheet.write('B1','Crypto Currency',bold)
worksheet.write('C1','Open',bold)
worksheet.write('D1', 'High', bold)
worksheet.write('E1', 'Low', bold)
worksheet.write('F1', 'Close', bold)
worksheet.write('G1', 'Volume', bold)
worksheet.write('H1', 'Market Cap', bold)
worksheet.write('I1', 'SMA 50', bold)
worksheet.write('J1', 'SMA 200', bold)
worksheet.write('K1', 'EMA 50', bold)
worksheet.write('L1', 'EMA 200', bold)
worksheet.write('M1', 'Close Change', bold)
worksheet.write('N1', 'Gain', bold)
worksheet.write('O1', 'Loss', bold)
worksheet.write('P1', 'Average Gain', bold)
worksheet.write('Q1', 'Average Loss', bold)
worksheet.write('R1', 'RS', bold)
worksheet.write('S1', '14-day RSI', bold)


wb = openpyxl.load_workbook('/Users/mohammadmuntasir/Downloads/CoinMarket.xlsx')

sheet1 = wb.get_sheet_by_name('bitcoin')
sheet2 = wb.get_sheet_by_name('ripple')
sheet3 = wb.get_sheet_by_name('ethereum')
sheet4 = wb.get_sheet_by_name('bitcoincash')

formatd = workbook.add_format({'num_format': 'mmm d yyyy'})
count = 1

def writeData(sheet, name, count):
    dates = []
    openList = []
    highList = []
    lowList = []
    closeList = []
    volList = []
    mcList = []
    ssmaList = []
    lsmaList = []
    semaList = []
    lemaList = []

    for i in range (2,sheet.max_row+1):
        dates.append(sheet['A' + str(i)].value)
        openList.append(sheet['B' + str(i)].value)
        highList.append(sheet['C' + str(i)].value)
        lowList.append(sheet['D' + str(i)].value)
        closeList.append(sheet['E' + str(i)].value)
        volList.append(sheet['F' + str(i)].value)
        mcList.append(sheet['G' + str(i)].value)

    dates.reverse()
    openList.reverse()
    highList.reverse()
    lowList.reverse()
    closeList.reverse()
    volList.reverse()
    mcList.reverse()

    ssmaList, lsmaList, semaList, lemaList = infCalc(closeList)
    changeList, gains, losses, aveGainList, aveLossList, rsList, rsiList = rsiCalc(closeList)

    idx = 0

    for i in range (2,sheet.max_row+1):
        worksheet.write(count,0, dates[idx], formatd)
        worksheet.write(count,1, name)
        worksheet.write(count,2, openList[idx])
        worksheet.write(count,3, highList[idx])
        worksheet.write(count,4, lowList[idx])
        worksheet.write(count,5, closeList[idx])
        worksheet.write(count,6, volList[idx])
        worksheet.write(count,7, mcList[idx])
        worksheet.write(count,8, ssmaList[idx])
        worksheet.write(count,9, lsmaList[idx])
        worksheet.write(count,10, semaList[idx])
        worksheet.write(count,11, lemaList[idx])
        worksheet.write(count,12, changeList[idx])
        if (gains[idx] == 0):
            worksheet.write(count,13, '-')
        else:
            worksheet.write(count,13, gains[idx])
        if (losses[idx] == 0):
            worksheet.write(count,14, '-')
        else:
            worksheet.write(count,14, -1 * losses[idx])
        worksheet.write(count,15, aveGainList[idx])
        worksheet.write(count,16, aveLossList[idx])
        worksheet.write(count,17, rsList[idx])
        worksheet.write(count,18, rsiList[idx])
        idx = idx + 1
        count = count + 1
    return count

count = writeData(sheet1, 'Bitcoin', count)
count = writeData(sheet2, 'Ripple',count)
count = writeData(sheet3, 'Ethereum',count)
count = writeData(sheet4, 'Bitcoin Cash',count)


workbook.close()
end = time.time()
elapsed = end - start
print("This program took " + str(round(elapsed,2)) + " seconds to run")
